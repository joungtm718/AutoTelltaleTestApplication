import tkinter as tk
from tkinter import filedialog, messagebox
import os
import openpyxl
import cantools
import can
from openpyxl.styles import PatternFill, Font

def get_cell_fill(color_value):
    return PatternFill(start_color=color_value,
                       end_color=color_value,
                       fill_type='solid')

def send_TT_msg(dbc, sheet, log_file_path=None):
    """Sends a single message and optionally logs CAN messages."""
    with can.Bus(interface='vector', app_name='CANalyzer', channel=0, bitrate=500000) as bus:
        if log_file_path:
            logger = can.Logger(log_file_path)
        else:
            logger = None

        for row in range(2, sheet.max_row + 1):
            print(f"\nTest Case {row - 1}/{sheet.max_row - 1}:")
            if sheet.cell(row=row, column=3).value is not None:
                message = sheet.cell(row=row, column=3).value
                signal = sheet.cell(row=row, column=4).value
                can_value = sheet.cell(row=row, column=5).value
                sheet.cell(row=row, column=6).value = "Invalid Case"
                sheet.cell(row=row, column=6).fill = get_cell_fill('D3D3D3')  # invalid fill
                print(f"Check {sheet.cell(row=row, column=1).value} TT in {sheet.cell(row=row, column=2).value} state with {message, signal, can_value}")
                try:
                    TT_message = dbc.get_message_by_name(message)
                    signals = {}
                    sig_status = False
                    for i in TT_message.signals:
                        if signal == i.name:
                            signals[i.name] = int(can_value, 16)
                            sig_status = True
                        else:
                            signals[i.name] = i.initial if i.initial is not None else 0  # if there is no initial value, set as 0

                    if sig_status:
                        try:
                            if TT_message.name == "CGW_PC2":
                                data = [0x0, 0x4, 0x0, 0x0, 0x0, 0x0, 0x0, 0x0] if can_value == '0x1' else [0x0, 0x0, 0x0, 0x0, 0x0, 0x0, 0x0, 0x0]
                            elif(TT_message.name == "EMS12") :
                                if can_value == '0xE1' :
                                    data = [0x0,0xE1,0x0,0x0,0x0,0x0,0x0,0x0]
                                elif can_value == '0xDD' : 
                                    data = [0x0,0xDD,0x0,0x0,0x0,0x0,0x0,0x0]
                                elif can_value == '0xFF' : 
                                    data = [0x0,0xFF,0x0,0x0,0x0,0x0,0x0,0x0]
                                else : 
                                    data = TT_message.encode(signals)
                            else:
                                data = TT_message.encode(signals)

                            msg = can.Message(arbitration_id=int(TT_message.frame_id), data=data, is_extended_id=False)
                            msg_timeout = int(TT_message.cycle_time) / 1000
                            task = bus.send_periodic(msg, msg_timeout)
                            try:
                                Result = input("Enter TT status as per case (Y/N):")
                                if Result.lower() == 'y':
                                    sheet.cell(row=row, column=6).value = "Pass"
                                    sheet.cell(row=row, column=6).fill = get_cell_fill('FFC6EFCE')  # valid fill
                                    sheet.cell(row=row, column=6).font = Font(color='006100')  # Green text
                                elif Result.lower() == 'n':
                                    sheet.cell(row=row, column=6).value = "Fail"
                                    sheet.cell(row=row, column=6).fill = get_cell_fill('FFFFC7CE')  # invalid fill
                                    sheet.cell(row=row, column=6).font = Font(color='9C0006')  # Red text
                                else:
                                    print("No response")
                                if logger:
                                    logger.on_message_received(msg)
                            finally:
                                task.stop()
                        except can.CanError:
                            print("Message NOT sent")
                    else:
                        print(f"{signal} not present in {message}")
                except KeyError:
                    print(f"{message} not present in DBC")
            else:
                print("Invalid Test case in FTP")

# Function to run CAN test
def send_TT_msg_gui():
    ftp_path = ftp_file.get()
    dbc_path = dbc_file.get()

    if not ftp_path or not dbc_path:
        messagebox.showerror("Error", "Both FTP Excel and DBC files must be selected!")
        return

    log_file_path = None
    if log_checkbox_var.get():
        log_file_path = log_file_entry.get()
        if not log_file_path:
            messagebox.showerror("Error", "Log file path must be selected!")
            return

    try:
        # Load Excel and DBC files
        wb_obj = openpyxl.load_workbook(ftp_path)
        ftp_sheet = wb_obj.active
        dbc = cantools.database.load_file(dbc_path)

        # Perform CAN test and log messages if needed
        send_TT_msg(dbc, ftp_sheet, log_file_path)

        wb_obj.save(ftp_path)
        messagebox.showinfo("Success", "CAN test completed and results saved!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

# Function to select log file path
def select_log_file():
    log_file_path = filedialog.asksaveasfilename(defaultextension=".asc", filetypes=[("ASC Files", "*.asc")])
    if log_file_path:
        log_file_entry.delete(0, tk.END)
        log_file_entry.insert(0, log_file_path)

# File selector helper
def select_file(entry_field, file_type):
    filetypes = [("Excel Files", "*.xlsx")] if file_type == "Excel" else [("DBC Files", "*.dbc")]
    filename = filedialog.askopenfilename(title=f"Select {file_type} File", filetypes=filetypes)
    if filename:
        entry_field.delete(0, tk.END)
        entry_field.insert(0, filename)

if __name__ == "__main__":
    # Tkinter GUI
    root = tk.Tk()
    root.title("Visteon Function Test for Cluster Ver. 0.2")

    # FTP Excel File
    tk.Label(root, text="FTP Excel File:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
    ftp_file = tk.Entry(root, width=40)
    ftp_file.grid(row=0, column=1, padx=10, pady=5)
    tk.Button(root, text="Browse", command=lambda: select_file(ftp_file, "Excel")).grid(row=0, column=2, padx=10, pady=5)

    # DBC File
    tk.Label(root, text="DBC File:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
    dbc_file = tk.Entry(root, width=40)
    dbc_file.grid(row=1, column=1, padx=10, pady=5)
    tk.Button(root, text="Browse", command=lambda: select_file(dbc_file, "DBC")).grid(row=1, column=2, padx=10, pady=5)

    # CAN Log Configuration
    log_checkbox_var = tk.BooleanVar()
    tk.Checkbutton(root, text="Log CAN messages", variable=log_checkbox_var).grid(row=2, column=0, columnspan=3, pady=5)
    tk.Label(root, text="Log File:").grid(row=3, column=0, padx=10, pady=5, sticky="w")
    log_file_entry = tk.Entry(root, width=40)
    log_file_entry.grid(row=3, column=1, padx=10, pady=5)
    tk.Button(root, text="Save As...", command=select_log_file).grid(row=3, column=2, padx=10, pady=5)

    # Buttons
    tk.Button(root, text="Run CAN Test", command=send_TT_msg_gui, bg="lightblue").grid(row=4, column=0, columnspan=3, pady=10)

    # Run the Tkinter main loop
    root.mainloop()